"""
Office 文档预览转换：DOCX/XLSX -> HTML
"""
from __future__ import annotations

import io
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, time
from html import escape
from typing import Iterator, Tuple

from docx import Document as DocxDocument
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph as DocxParagraph
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MAX_XLSX_PREVIEW_ROWS = 200
MAX_XLSX_PREVIEW_COLS = 60


class PreviewConversionError(Exception):
    """预览转换失败。"""


def _sheet_anchor_id(name: str) -> str:
    token = re.sub(r"[^0-9a-zA-Z_-]+", "-", name or "sheet").strip("-")
    return token or "sheet"


def is_office_preview_supported(mime_type: str, filename: str = "") -> bool:
    normalized = (mime_type or "").split(";", 1)[0].strip().lower()
    lower_name = (filename or "").lower()
    return normalized in {DOCX_MIME, XLSX_MIME} or lower_name.endswith((".docx", ".xlsx"))


def convert_office_to_html(file_content: bytes, mime_type: str, filename: str = "") -> bytes:
    normalized = (mime_type or "").split(";", 1)[0].strip().lower()
    lower_name = (filename or "").lower()

    try:
        if normalized == DOCX_MIME or lower_name.endswith(".docx"):
            html_text = _convert_docx_to_html(file_content, filename)
        elif normalized == XLSX_MIME or lower_name.endswith(".xlsx"):
            html_text = _convert_xlsx_to_html(file_content, filename)
        else:
            raise PreviewConversionError(f"不支持的Office预览类型: mime={mime_type}, filename={filename}")
    except PreviewConversionError:
        raise
    except Exception as exc:
        raise PreviewConversionError(str(exc)) from exc

    return html_text.encode("utf-8")


def _build_html_document(title: str, subtitle: str, body: str) -> str:
    safe_title = escape(title or "文档预览")
    safe_subtitle = escape(subtitle or "")
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{safe_title}</title>
  <style>
    :root {{
      color-scheme: light;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      padding: 24px;
      font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      color: #19231d;
      background:
        radial-gradient(circle at 15% 12%, rgba(31, 111, 95, 0.09), transparent 28%),
        radial-gradient(circle at 86% 6%, rgba(217, 167, 104, 0.16), transparent 24%),
        linear-gradient(180deg, #f6f8f7 0%, #eef3f1 100%);
    }}
    .container {{
      max-width: 1160px;
      margin: 0 auto;
      background: rgba(255, 255, 255, 0.9);
      border: 1px solid rgba(31, 111, 95, 0.14);
      border-radius: 16px;
      box-shadow: 0 18px 44px rgba(31, 42, 36, 0.08);
      overflow: hidden;
    }}
    .header {{
      padding: 18px 22px;
      background: linear-gradient(120deg, #1f6f5f 0%, #2c8e78 100%);
      color: #f4f8f6;
    }}
    .header h1 {{
      margin: 0;
      font-size: 18px;
      line-height: 1.35;
      font-weight: 700;
    }}
    .header p {{
      margin: 6px 0 0;
      opacity: 0.92;
      font-size: 13px;
    }}
    .content {{
      padding: 20px 22px 24px;
    }}
    .docx-block p {{
      margin: 0 0 10px;
      white-space: pre-wrap;
      line-height: 1.75;
      word-break: break-word;
    }}
    .docx-block h1, .docx-block h2, .docx-block h3, .docx-block h4 {{
      margin: 14px 0 8px;
      color: #1a322a;
      line-height: 1.45;
    }}
    .docx-block h1 {{ font-size: 24px; }}
    .docx-block h2 {{ font-size: 21px; }}
    .docx-block h3 {{ font-size: 18px; }}
    .docx-block h4 {{ font-size: 16px; }}
    .docx-block .list-item {{
      padding-left: 18px;
      text-indent: -14px;
    }}
    .docx-block table, .xlsx-sheet table {{
      width: 100%;
      border-collapse: collapse;
      margin: 12px 0 16px;
      table-layout: fixed;
    }}
    .docx-block th, .docx-block td, .xlsx-sheet th, .xlsx-sheet td {{
      border: 1px solid #d6dde4;
      padding: 6px 8px;
      font-size: 13px;
      line-height: 1.6;
      vertical-align: top;
      word-break: break-word;
      background: rgba(255, 255, 255, 0.94);
    }}
    .docx-block th {{
      background: #f0f5f3;
      font-weight: 600;
    }}
    .sheet-anchor-list {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 0 0 14px;
      padding: 0;
      list-style: none;
    }}
    .sheet-anchor-list a {{
      display: inline-block;
      padding: 4px 10px;
      border-radius: 999px;
      font-size: 12px;
      color: #1f6f5f;
      background: rgba(31, 111, 95, 0.1);
      text-decoration: none;
    }}
    .sheet-anchor-list a:hover {{
      background: rgba(31, 111, 95, 0.18);
    }}
    .xlsx-sheet {{
      margin-bottom: 20px;
      border: 1px solid #d6dde4;
      border-radius: 12px;
      overflow: hidden;
      background: #fff;
    }}
    .xlsx-sheet__title {{
      margin: 0;
      padding: 10px 12px;
      font-size: 14px;
      font-weight: 700;
      color: #1a322a;
      background: #eff5f3;
      border-bottom: 1px solid #d6dde4;
    }}
    .xlsx-sheet__wrap {{
      overflow-x: auto;
    }}
    .xlsx-sheet thead th {{
      background: #f6f8fb;
      position: sticky;
      top: 0;
      z-index: 2;
      text-align: center;
      font-weight: 600;
    }}
    .xlsx-row-head {{
      background: #f6f8fb;
      text-align: center;
      width: 56px;
      min-width: 56px;
      position: sticky;
      left: 0;
      z-index: 1;
      font-weight: 600;
    }}
    .xlsx-note {{
      margin: 0;
      padding: 10px 12px;
      font-size: 12px;
      color: #6b7b72;
      border-top: 1px solid #e2e8ee;
      background: #fafcfd;
    }}
    .empty-tip {{
      color: #71847a;
      font-size: 14px;
      margin: 4px 0;
    }}
    @media (max-width: 920px) {{
      body {{ padding: 14px; }}
      .container {{ border-radius: 12px; }}
      .content {{ padding: 14px; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header class="header">
      <h1>{safe_title}</h1>
      <p>{safe_subtitle}</p>
    </header>
    <main class="content">{body}</main>
  </div>
</body>
</html>"""


def _iter_docx_blocks(doc: DocxDocument) -> Iterator[object]:
    for child in doc.element.body.iterchildren():
        if child.tag.endswith("}p"):
            yield DocxParagraph(child, doc)
        elif child.tag.endswith("}tbl"):
            yield DocxTable(child, doc)


def _render_docx_paragraph(paragraph: DocxParagraph) -> str:
    text_parts = []
    for run in paragraph.runs:
        run_text = escape(run.text or "").replace("\n", "<br/>")
        if not run_text:
            continue
        if run.bold:
            run_text = f"<strong>{run_text}</strong>"
        if run.italic:
            run_text = f"<em>{run_text}</em>"
        if run.underline:
            run_text = f"<u>{run_text}</u>"
        text_parts.append(run_text)

    content = "".join(text_parts).strip()
    if not content:
        fallback = escape(paragraph.text or "").replace("\n", "<br/>").strip()
        content = fallback
    if not content:
        return ""

    style_name = str(getattr(paragraph.style, "name", "") or "").lower()
    tag = "p"
    if "heading" in style_name:
        level = 2
        for ch in style_name:
            if ch.isdigit():
                level = max(1, min(4, int(ch)))
                break
        tag = f"h{level}"
    elif style_name.strip() == "title":
        tag = "h1"

    extra_class = ""
    if "list bullet" in style_name:
        content = f"&#8226; {content}"
        extra_class = " class=\"list-item\""
    elif "list number" in style_name:
        extra_class = " class=\"list-item\""

    return f"<{tag}{extra_class}>{content}</{tag}>"


def _render_docx_table(table: DocxTable) -> str:
    rows = []
    for row in table.rows:
        cells = [escape(cell.text or "").replace("\n", "<br/>") for cell in row.cells]
        rows.append(cells)

    if not rows:
        return ""

    has_header = len(rows) > 1 and any(cell.strip() for cell in rows[0])
    html_rows = []
    for idx, row in enumerate(rows):
        tag = "th" if has_header and idx == 0 else "td"
        cols = "".join([f"<{tag}>{cell or '&nbsp;'}</{tag}>" for cell in row])
        html_rows.append(f"<tr>{cols}</tr>")

    return f"<table><tbody>{''.join(html_rows)}</tbody></table>"


def _convert_docx_to_html(file_content: bytes, filename: str) -> str:
    try:
        doc = DocxDocument(io.BytesIO(file_content))
    except Exception as exc:
        raise PreviewConversionError(f"DOCX 转换失败: {str(exc)}") from exc

    blocks = []
    for block in _iter_docx_blocks(doc):
        if isinstance(block, DocxParagraph):
            html_block = _render_docx_paragraph(block)
            if html_block:
                blocks.append(html_block)
        elif isinstance(block, DocxTable):
            html_table = _render_docx_table(block)
            if html_table:
                blocks.append(html_table)

    if not blocks:
        blocks.append('<p class="empty-tip">文档无可展示内容。</p>')

    title = filename or "DOCX 文档预览"
    subtitle = "服务端转换为 HTML 预览（尽量保留标题、段落与表格结构）"
    body = f"<section class=\"docx-block\">{''.join(blocks)}</section>"
    return _build_html_document(title, subtitle, body)


def _xlsx_color_to_css(color) -> str | None:
    if not color:
        return None

    value = getattr(color, "rgb", None)
    color_type = getattr(color, "type", None)
    if color_type != "rgb" or not value:
        return None

    text = str(value)
    if len(text) == 8:
        text = text[2:]
    if len(text) != 6:
        return None
    return f"#{text.lower()}"


def _xlsx_cell_style(cell: Cell) -> str:
    styles = []

    font = cell.font
    if font:
        if font.bold:
            styles.append("font-weight: 700")
        if font.italic:
            styles.append("font-style: italic")
        text_color = _xlsx_color_to_css(font.color)
        if text_color and text_color != "#000000":
            styles.append(f"color: {text_color}")

    fill = cell.fill
    if fill and getattr(fill, "patternType", None) not in (None, "none"):
        bg_color = _xlsx_color_to_css(getattr(fill, "fgColor", None))
        if bg_color and bg_color != "#000000":
            styles.append(f"background-color: {bg_color}")

    alignment = cell.alignment
    if alignment:
        if alignment.horizontal in {"left", "center", "right", "justify"}:
            styles.append(f"text-align: {alignment.horizontal}")
        if alignment.vertical in {"top", "center", "bottom"}:
            styles.append(f"vertical-align: {alignment.vertical}")
        if alignment.wrap_text:
            styles.append("white-space: pre-wrap")

    border = cell.border
    if border and any([
        getattr(border.left, "style", None),
        getattr(border.right, "style", None),
        getattr(border.top, "style", None),
        getattr(border.bottom, "style", None),
    ]):
        styles.append("border: 1px solid #9aa8b4")

    return "; ".join(styles)


def _format_xlsx_value(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, date):
        text = value.strftime("%Y-%m-%d")
    elif isinstance(value, time):
        text = value.strftime("%H:%M:%S")
    else:
        text = str(value)
    return escape(text).replace("\n", "<br/>")


def _merged_cell_maps(ws) -> Tuple[dict[Tuple[int, int], Tuple[int, int]], set[Tuple[int, int]]]:
    start_map: dict[Tuple[int, int], Tuple[int, int]] = {}
    covered: set[Tuple[int, int]] = set()
    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        start_map[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    covered.add((r, c))
    return start_map, covered


def _extract_xml_ns(tag: str, default_ns: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag[1: tag.find("}")]
    return default_ns


def _column_index_from_ref(cell_ref: str) -> int:
    letters = "".join(ch for ch in (cell_ref or "") if ch.isalpha())
    if not letters:
        return 0

    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def _safe_xl_path(target: str) -> str:
    text = str(target or "").replace("\\", "/")
    if text.startswith("/"):
        text = text[1:]
    if not text.startswith("xl/"):
        text = f"xl/{text.lstrip('./')}"
    return text


def _read_xlsx_cell_value(cell_node, ns: dict[str, str], shared_strings: list[str]) -> str:
    cell_type = cell_node.get("t")
    v_node = cell_node.find("x:v", ns)
    inline_nodes = cell_node.findall(".//x:is//x:t", ns)
    formula_node = cell_node.find("x:f", ns)

    if cell_type == "s" and v_node is not None and v_node.text:
        try:
            idx = int(v_node.text)
            return shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
        except Exception:
            return v_node.text or ""

    if cell_type == "inlineStr":
        if inline_nodes:
            return "".join((n.text or "") for n in inline_nodes)
        return ""

    if cell_type == "b" and v_node is not None and v_node.text is not None:
        return "TRUE" if str(v_node.text).strip() == "1" else "FALSE"

    if v_node is not None and v_node.text is not None:
        return str(v_node.text)

    if formula_node is not None and formula_node.text:
        return f"={formula_node.text}"

    return ""


def _extract_xlsx_rows_from_xml(file_content: bytes) -> list[dict[str, object]]:
    sheets: list[dict[str, object]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_content)) as zf:
            names = set(zf.namelist())

            shared_strings: list[str] = []
            if "xl/sharedStrings.xml" in names:
                ss_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                ss_ns = {"x": _extract_xml_ns(ss_root.tag, "http://schemas.openxmlformats.org/spreadsheetml/2006/main")}
                for si in ss_root.findall("x:si", ss_ns):
                    text_parts = [node.text or "" for node in si.findall(".//x:t", ss_ns)]
                    shared_strings.append("".join(text_parts))

            sheet_sources: list[tuple[str, str]] = []
            if "xl/workbook.xml" in names:
                wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
                wb_ns = {"x": _extract_xml_ns(wb_root.tag, "http://schemas.openxmlformats.org/spreadsheetml/2006/main")}

                rel_map: dict[str, str] = {}
                if "xl/_rels/workbook.xml.rels" in names:
                    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                    for rel in rel_root.iter():
                        if not rel.tag.endswith("Relationship"):
                            continue
                        rid = rel.attrib.get("Id")
                        target = rel.attrib.get("Target")
                        if rid and target:
                            rel_map[rid] = target

                rel_id_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                for sheet_node in wb_root.findall(".//x:sheets/x:sheet", wb_ns):
                    name = sheet_node.attrib.get("name") or f"Sheet{len(sheet_sources) + 1}"
                    rid = sheet_node.attrib.get(rel_id_attr)
                    target = rel_map.get(rid or "")
                    if not target:
                        continue
                    normalized_path = _safe_xl_path(target)
                    if "worksheets/" not in normalized_path:
                        continue
                    if normalized_path in names:
                        sheet_sources.append((name, normalized_path))

            if not sheet_sources:
                fallback_sheet_files = sorted(
                    name for name in names
                    if name.startswith("xl/worksheets/") and name.endswith(".xml")
                )
                for idx, path in enumerate(fallback_sheet_files, start=1):
                    sheet_sources.append((f"Sheet{idx}", path))

            for sheet_name, sheet_path in sheet_sources:
                try:
                    root = ET.fromstring(zf.read(sheet_path))
                except Exception:
                    continue

                sheet_ns = {"x": _extract_xml_ns(root.tag, "http://schemas.openxmlformats.org/spreadsheetml/2006/main")}
                parsed_rows: list[list[str]] = []

                for row_node in root.findall(".//x:sheetData/x:row", sheet_ns):
                    row_value_map: dict[int, str] = {}
                    last_col = 0
                    for cell_node in row_node.findall("x:c", sheet_ns):
                        col_idx = _column_index_from_ref(cell_node.get("r", ""))
                        if col_idx <= 0:
                            col_idx = last_col + 1

                        value = _read_xlsx_cell_value(cell_node, sheet_ns, shared_strings).strip()
                        if value:
                            row_value_map[col_idx] = value
                        last_col = max(last_col, col_idx)

                    if not row_value_map:
                        continue

                    max_col = max(row_value_map.keys())
                    parsed_rows.append([row_value_map.get(i, "") for i in range(1, max_col + 1)])

                if parsed_rows:
                    sheets.append({"title": sheet_name, "rows": parsed_rows})
    except Exception:
        return []

    return sheets


def _render_xlsx_fallback_sheet(sheet_name: str, rows: list[list[str]]) -> str:
    max_row = max(1, len(rows))
    max_col = max((len(r) for r in rows), default=1)
    render_rows = min(max_row, MAX_XLSX_PREVIEW_ROWS)
    render_cols = min(max_col, MAX_XLSX_PREVIEW_COLS)

    head_cols = "".join([f"<th>{escape(get_column_letter(col))}</th>" for col in range(1, render_cols + 1)])
    table_parts = [
        '<div class="xlsx-sheet__wrap">',
        '<table>',
        f'<thead><tr><th class="xlsx-row-head">#</th>{head_cols}</tr></thead>',
        '<tbody>',
    ]

    for row_idx in range(1, render_rows + 1):
        table_parts.append(f'<tr><th class="xlsx-row-head">{row_idx}</th>')
        row_values = rows[row_idx - 1] if row_idx - 1 < len(rows) else []
        for col_idx in range(1, render_cols + 1):
            raw_value = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else ""
            value = escape(str(raw_value or "")).replace("\n", "<br/>")
            table_parts.append(f"<td>{value or '&nbsp;'}</td>")
        table_parts.append("</tr>")

    table_parts.extend(["</tbody>", "</table>", "</div>"])

    note_parts = ["兼容模式：工作簿结构异常，已按XML回退解析"]
    if max_row > render_rows:
        note_parts.append(f"行数过多，已截断为前 {render_rows} 行")
    if max_col > render_cols:
        note_parts.append(f"列数过多，已截断为前 {render_cols} 列")
    note_html = f'<p class="xlsx-note">{"；".join(note_parts)}</p>'

    display_name = escape(sheet_name or "Sheet")
    anchor = _sheet_anchor_id(sheet_name or "sheet")
    return (
        f'<section class="xlsx-sheet" id="sheet-{anchor}">'
        f'<h3 class="xlsx-sheet__title">{display_name}</h3>'
        f'{"".join(table_parts)}{note_html}'
        '</section>'
    )


def _render_xlsx_sheet(ws) -> str:
    max_row = max(1, int(ws.max_row or 1))
    max_col = max(1, int(ws.max_column or 1))
    render_rows = min(max_row, MAX_XLSX_PREVIEW_ROWS)
    render_cols = min(max_col, MAX_XLSX_PREVIEW_COLS)

    merged_start, merged_covered = _merged_cell_maps(ws)

    head_cols = "".join([f"<th>{escape(get_column_letter(col))}</th>" for col in range(1, render_cols + 1)])
    table_parts = [
        '<div class="xlsx-sheet__wrap">',
        '<table>',
        f'<thead><tr><th class="xlsx-row-head">#</th>{head_cols}</tr></thead>',
        '<tbody>',
    ]

    for row_idx in range(1, render_rows + 1):
        table_parts.append(f'<tr><th class="xlsx-row-head">{row_idx}</th>')
        for col_idx in range(1, render_cols + 1):
            if (row_idx, col_idx) in merged_covered:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            value = _format_xlsx_value(cell)
            style = _xlsx_cell_style(cell)

            attrs = ""
            if (row_idx, col_idx) in merged_start:
                rowspan, colspan = merged_start[(row_idx, col_idx)]
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
            if style:
                attrs += f' style="{style}"'

            table_parts.append(f"<td{attrs}>{value or '&nbsp;'}</td>")
        table_parts.append("</tr>")

    table_parts.extend(["</tbody>", "</table>", "</div>"])

    note_parts = []
    if max_row > render_rows:
        note_parts.append(f"行数过多，已截断为前 {render_rows} 行")
    if max_col > render_cols:
        note_parts.append(f"列数过多，已截断为前 {render_cols} 列")
    note_html = f'<p class="xlsx-note">{"；".join(note_parts)}</p>' if note_parts else ""

    sheet_name = escape(ws.title or "Sheet")
    anchor = _sheet_anchor_id(ws.title or "sheet")
    return (
        f'<section class="xlsx-sheet" id="sheet-{anchor}">'
        f'<h3 class="xlsx-sheet__title">{sheet_name}</h3>'
        f'{"".join(table_parts)}{note_html}'
        '</section>'
    )


def _convert_xlsx_to_html(file_content: bytes, filename: str) -> str:
    workbook = None
    workbook_error: Exception | None = None
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)
    except Exception as exc:
        workbook_error = exc

    try:
        sheets = workbook.worksheets if workbook else []
        if sheets:
            anchor_items = []
            for sheet in sheets:
                name = escape(sheet.title or "Sheet")
                anchor = _sheet_anchor_id(sheet.title or "sheet")
                anchor_items.append(f'<li><a href="#sheet-{anchor}">{name}</a></li>')

            section_html = "".join([_render_xlsx_sheet(sheet) for sheet in sheets])
            body = (
                f'<ul class="sheet-anchor-list">{"".join(anchor_items)}</ul>'
                f'{section_html}'
            )
            title = filename or "XLSX 文档预览"
            subtitle = "服务端转换为 HTML 预览（尽量保留工作表结构与单元格样式）"
            return _build_html_document(title, subtitle, body)

        # 兼容回退：openpyxl 未拿到可展示工作表时，直接按 XML 解析 worksheet 数据。
        fallback_sheets = _extract_xlsx_rows_from_xml(file_content)
        if fallback_sheets:
            anchor_items = []
            for sheet in fallback_sheets:
                name = str(sheet.get("title") or "Sheet")
                safe_name = escape(name)
                anchor = _sheet_anchor_id(name)
                anchor_items.append(f'<li><a href="#sheet-{anchor}">{safe_name}</a></li>')

            section_html = "".join([
                _render_xlsx_fallback_sheet(
                    str(sheet.get("title") or "Sheet"),
                    sheet.get("rows") if isinstance(sheet.get("rows"), list) else [],
                )
                for sheet in fallback_sheets
            ])
            body = (
                f'<ul class="sheet-anchor-list">{"".join(anchor_items)}</ul>'
                f'{section_html}'
            )
            title = filename or "XLSX 文档预览"
            subtitle = "服务端转换为 HTML 预览（兼容模式）"
            return _build_html_document(title, subtitle, body)

        if workbook_error is not None:
            raise PreviewConversionError(f"XLSX 转换失败: {str(workbook_error)}") from workbook_error

        body = '<p class="empty-tip">工作簿中没有可展示的工作表。</p>'
        return _build_html_document(filename or "XLSX 文档预览", "服务端转换为 HTML 预览", body)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
