"""
模板服务 - 模板 CRUD、版本管理、提示词生成
"""
import asyncio
import uuid
import json
import ast
import csv
import io
import re
import logging
from datetime import datetime
from typing import Optional, List, Tuple, Dict, Any, Callable, Awaitable
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, delete
from sqlalchemy.orm import selectinload
from app.models.template import Template, TemplateField, TemplateVersion, TemplateCategory, TemplateStatus, FieldType
from app.models.document import Document, DocumentStatus
from app.models.system import LLMConfig as LLMConfigModel, SystemConfig
from app.schemas.template import TemplateCreate, TemplateUpdate, TemplateCategoryCreate
from app.core.exceptions import NotFoundException, ValidationException
from app.llm.base_adapter import LLMMessage
from app.llm.factory import get_default_adapter, get_default_llm_config, create_adapter_from_db_config
from app.core.storage import storage
from app.config import settings
from app.processors.factory import get_processor
from app.processors.mime_resolver import normalize_mime_type


TEMPLATE_LLM_LOG_MAX_CHARS = 4000
TEMPLATE_INFER_CHUNK_SIZE = 3000
TEMPLATE_INFER_CHUNK_OVERLAP = 500
TEMPLATE_INFER_MAX_CHUNKS = 6
TEMPLATE_INFER_CHUNK_MAX_FIELDS = 8
TEMPLATE_INFER_VERIFY_TEXT_LIMIT = 12000
TEMPLATE_INFER_CHUNK_CONCURRENCY = 3
logger = logging.getLogger(__name__)


TEMPLATE_TABULAR_HEADERS: List[Tuple[str, str]] = [
    ("字段标识", "field_name"),
    ("显示名称", "field_display_name"),
    ("字段类型", "field_type"),
    ("是否必填", "field_required"),
    ("字段描述", "field_description"),
    ("提取提示", "field_extraction_hints"),
    ("排序", "field_sort_order"),
]

TEMPLATE_TABULAR_ALIAS: Dict[str, List[str]] = {
    "template_name": ["template_name", "template", "模板名称", "模板名"],
    "template_description": ["template_description", "模板描述", "描述"],
    "template_status": ["template_status", "status", "模板状态"],
    "field_name": ["field_name", "name", "字段标识", "字段名", "标识"],
    "field_display_name": ["field_display_name", "display_name", "显示名称", "字段显示名", "名称"],
    "field_type": ["field_type", "type", "字段类型", "类型"],
    "field_required": ["field_required", "required", "是否必填", "必填"],
    "field_description": ["field_description", "description", "字段描述", "字段说明"],
    "field_extraction_hints": ["field_extraction_hints", "extraction_hints", "提取提示", "提取说明"],
    "field_sort_order": ["field_sort_order", "sort_order", "排序", "顺序"],
}


class TemplateService:
    def __init__(self, db: AsyncSession):
        self.db = db

    def _truncate_log_text(self, value: Any, max_chars: int = TEMPLATE_LLM_LOG_MAX_CHARS) -> str:
        if isinstance(value, str):
            text = value
        else:
            try:
                text = json.dumps(value, ensure_ascii=False)
            except Exception:
                text = str(value)

        if len(text) <= max_chars:
            return text
        return f"{text[:max_chars]}... [truncated {len(text) - max_chars} chars]"

    async def _resolve_infer_adapter_and_config(self):
        """模板自动提取优先使用系统默认 LLM 配置，回退到环境默认。"""
        llm_config = get_default_llm_config()

        try:
            sc_result = await self.db.execute(
                select(SystemConfig).where(SystemConfig.key == "default_llm_config_id")
            )
            sc = sc_result.scalar_one_or_none()
            default_cfg_id = sc.value if sc and sc.value else None

            if default_cfg_id:
                cfg_result = await self.db.execute(
                    select(LLMConfigModel).where(
                        LLMConfigModel.id == default_cfg_id,
                        LLMConfigModel.is_active == True,
                    )
                )
                cfg = cfg_result.scalar_one_or_none()
                if cfg and cfg.api_key_encrypted and cfg.base_url:
                    llm_config.model = cfg.model_name or llm_config.model
                    adapter = create_adapter_from_db_config(
                        cfg.api_key_encrypted,
                        cfg.base_url,
                        model=llm_config.model,
                    )
                    return adapter, llm_config
        except Exception as e:
            logger.warning("模板推断加载DB LLM配置失败，回退环境默认配置: error=%s", e)

        return get_default_adapter(), llm_config

    async def create(self, data: TemplateCreate, creator_id: str) -> Template:
        """创建模板"""
        template = Template(
            id=str(uuid.uuid4()),
            name=data.name,
            description=data.description,
            category_id=data.category_id,
            creator_id=creator_id,
            is_public=data.is_public,
            tags=data.tags,
            document_types=data.document_types,
            system_prompt=data.system_prompt,
            extraction_prompt_template=data.extraction_prompt_template,
            few_shot_examples=data.few_shot_examples,
            status=data.status or TemplateStatus.DRAFT,
            current_version=1,
        )
        self.db.add(template)
        await self.db.flush()

        # 创建字段
        for idx, field_data in enumerate(data.fields):
            field = TemplateField(
                id=str(uuid.uuid4()),
                template_id=template.id,
                name=field_data.name,
                display_name=field_data.display_name,
                field_type=field_data.field_type,
                description=field_data.description,
                required=field_data.required,
                default_value=field_data.default_value,
                validation_rules=field_data.validation_rules,
                extraction_hints=field_data.extraction_hints,
                region_config=field_data.region_config,
                sort_order=field_data.sort_order or idx,
            )
            self.db.add(field)

        await self.db.flush()
        await self._save_version(template, "初始版本", creator_id)
        return template

    async def get_by_id(self, template_id: str) -> Template:
        # 显式预加载 fields，避免在序列化到 Pydantic 时触发延迟加载失败或字段为空
        result = await self.db.execute(
            select(Template).options(selectinload(Template.fields)).where(Template.id == template_id)
        )
        template = result.scalar_one_or_none()
        if not template:
            raise NotFoundException("模板", template_id)
        return template

    async def update(self, template: Template, data: TemplateUpdate, updater_id: str) -> Template:
        """更新模板（自动保存新版本）"""
        update_fields = data.model_dump(exclude_none=True, exclude={"change_description"})

        # 如果提交了 fields，则先处理字段替换（删除旧字段并新增）
        fields_data = update_fields.pop("fields", None)

        for field, value in update_fields.items():
            setattr(template, field, value)

        if fields_data is not None:
            # 删除已有字段
            await self.db.execute(delete(TemplateField).where(TemplateField.template_id == template.id))
            # 添加新字段
            for idx, field_data in enumerate(fields_data):
                # fields_data 是来自 pydantic.model_dump() 的 dict 列表，使用 dict 取值
                field = TemplateField(
                    id=str(uuid.uuid4()),
                    template_id=template.id,
                    name=field_data.get('name'),
                    display_name=field_data.get('display_name'),
                    field_type=field_data.get('field_type'),
                    description=field_data.get('description'),
                    required=field_data.get('required', False),
                    default_value=field_data.get('default_value'),
                    validation_rules=field_data.get('validation_rules') or {},
                    extraction_hints=field_data.get('extraction_hints'),
                    region_config=field_data.get('region_config') or {},
                    parent_field_id=field_data.get('parent_field_id'),
                    sort_order=field_data.get('sort_order', idx),
                )
                self.db.add(field)

        template.current_version += 1
        await self.db.flush()
        await self._save_version(template, data.change_description or "更新", updater_id)
        return template

    async def delete(self, template: Template):
        """软删除（归档）"""
        template.status = TemplateStatus.ARCHIVED
        await self.db.flush()

    async def list_templates(
        self,
        page: int = 1,
        page_size: int = 20,
        keyword: Optional[str] = None,
        category_id: Optional[str] = None,
        status: Optional[TemplateStatus] = None,
        is_public: Optional[bool] = None,
        creator_id: Optional[str] = None,
    ) -> Tuple[List[Template], int]:
        query = select(Template).where(Template.status != TemplateStatus.ARCHIVED)

        if keyword:
            query = query.where(
                or_(
                    Template.name.ilike(f"%{keyword}%"),
                    Template.description.ilike(f"%{keyword}%"),
                )
            )
        if category_id:
            query = query.where(Template.category_id == category_id)
        if status:
            query = query.where(Template.status == status)
        if is_public is not None:
            query = query.where(Template.is_public == is_public)
        if creator_id:
            query = query.where(Template.creator_id == creator_id)

        count_result = await self.db.execute(
            select(func.count()).select_from(query.subquery())
        )
        total = count_result.scalar()

        query = query.order_by(Template.updated_at.desc())
        query = query.offset((page - 1) * page_size).limit(page_size)
        result = await self.db.execute(query)
        return result.scalars().all(), total

    async def add_field(self, template: Template, field_data, updater_id: str) -> TemplateField:
        """向模板添加字段"""
        field = TemplateField(
            id=str(uuid.uuid4()),
            template_id=template.id,
            **field_data.model_dump(),
        )
        self.db.add(field)
        template.current_version += 1
        await self.db.flush()
        await self._save_version(template, f"添加字段: {field_data.display_name}", updater_id)
        return field

    async def _save_version(self, template: Template, description: str, creator_id: str):
        """保存当前版本快照

        为避免在非异步上下文触发 SQLAlchemy 的懒加载（从而导致 MissingGreenlet 错误），
        在这里显式在异步会话中重新查询并使用 `selectinload` 预加载 `fields` 关系。
        """
        # 在异步上下文中显式加载 fields，避免懒加载在后续同步上下文触发
        result = await self.db.execute(
            select(Template).options(selectinload(Template.fields)).where(Template.id == template.id)
        )
        tpl = result.scalar_one()
        snapshot = {
            "name": tpl.name,
            "description": tpl.description,
            "system_prompt": tpl.system_prompt,
            "extraction_prompt_template": tpl.extraction_prompt_template,
            "few_shot_examples": tpl.few_shot_examples,
            "fields": [
                {
                    "name": f.name,
                    "display_name": f.display_name,
                    "field_type": f.field_type.value if hasattr(f.field_type, 'value') else f.field_type,
                    "description": f.description,
                    "required": f.required,
                    "validation_rules": f.validation_rules,
                    "extraction_hints": f.extraction_hints,
                }
                for f in tpl.fields
            ],
        }
        version = TemplateVersion(
            id=str(uuid.uuid4()),
            template_id=template.id,
            version=template.current_version,
            template_snapshot=snapshot,
            change_description=description,
            created_by=creator_id,
        )
        self.db.add(version)

    # =====================
    # 模板分类管理
    # =====================
    async def create_category(self, data: TemplateCategoryCreate) -> TemplateCategory:
        category = TemplateCategory(
            id=str(uuid.uuid4()),
            name=data.name,
            description=data.description,
            parent_id=data.parent_id,
            sort_order=data.sort_order,
        )
        self.db.add(category)
        await self.db.flush()
        return category

    async def list_categories(self) -> List[TemplateCategory]:
        result = await self.db.execute(
            select(TemplateCategory).order_by(TemplateCategory.sort_order)
        )
        return result.scalars().all()

    async def import_template_from_file(
        self,
        file_content: bytes,
        filename: str,
        creator_id: str,
        template_name: Optional[str] = None,
        template_description: Optional[str] = None,
        template_status: Optional[TemplateStatus] = None,
        is_public: bool = False,
    ) -> Template:
        """从 CSV/XLSX 文件导入模板。"""
        ext = self._detect_template_file_ext(filename)
        headers, row_data = self._load_template_tabular_rows(file_content, ext)
        if not headers and not row_data:
            raise ValidationException("模板文件内容为空")

        normalized_rows = [self._normalize_template_import_row(row) for row in row_data] if row_data else []
        structured_mode = any(
            self._to_text(row.get("field_name")) or self._to_text(row.get("field_display_name"))
            for row in normalized_rows
        )

        resolved_name = (template_name or "").strip()
        if not resolved_name:
            if structured_mode:
                resolved_name = self._first_non_empty(normalized_rows, "template_name")
            if not resolved_name:
                resolved_name = self._derive_template_name_from_filename(filename)
        if not resolved_name:
            resolved_name = f"导入模板_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        resolved_desc = (template_description or "").strip() or (
            self._first_non_empty(normalized_rows, "template_description") if structured_mode else ""
        )
        resolved_status = template_status or (
            self._parse_template_status(self._first_non_empty(normalized_rows, "template_status"))
            if structured_mode
            else TemplateStatus.DRAFT
        )

        if structured_mode:
            fields = self._build_fields_from_structured_rows(normalized_rows)
        else:
            fields = self._build_fields_from_simple_headers(headers, row_data)

        if not fields:
            raise ValidationException("未识别到可导入字段，请检查文件表头")

        template_data = TemplateCreate(
            name=resolved_name[:128],
            description=resolved_desc or None,
            status=resolved_status,
            is_public=is_public,
            fields=fields,
        )
        return await self.create(template_data, creator_id)

    def export_template_file(self, template: Template, export_format: str) -> Tuple[bytes, str, str]:
        """导出模板为 CSV/XLSX 文件（仅字段表头）。"""
        fmt = (export_format or "xlsx").strip().lower()
        if fmt not in {"xlsx", "csv"}:
            raise ValidationException("仅支持导出 xlsx 或 csv 格式")

        export_headers = self._build_template_export_headers(template)
        filename = f"template_{template.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{fmt}"

        if fmt == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf)
            if export_headers:
                writer.writerow(export_headers)
            content = buf.getvalue().encode("utf-8-sig")
            return content, "text/csv", filename

        try:
            import openpyxl
        except ImportError as e:
            raise ValidationException(f"未安装 openpyxl，无法导出 xlsx: {e}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "字段模板"
        if export_headers:
            ws.append(export_headers)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename

    def _detect_template_file_ext(self, filename: str) -> str:
        ext = (filename or "").rsplit(".", 1)[-1].lower() if "." in (filename or "") else ""
        if ext not in {"xlsx", "csv"}:
            raise ValidationException("模板文件仅支持 .xlsx 或 .csv")
        return ext

    def _load_template_tabular_rows(self, file_content: bytes, ext: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        if ext == "csv":
            text = self._decode_csv_text(file_content)
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise ValidationException("CSV 文件缺少表头")
            headers = [self._to_text(h) for h in reader.fieldnames if self._to_text(h)]
            rows: List[Dict[str, Any]] = []
            for row in reader:
                if not isinstance(row, dict):
                    continue
                if any(self._to_text(v) for v in row.values()):
                    rows.append(row)
            return headers, rows

        try:
            import openpyxl
        except ImportError as e:
            raise ValidationException(f"未安装 openpyxl，无法读取 xlsx: {e}")

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return [], []

        headers = [self._to_text(cell) for cell in raw_rows[0]]
        if not any(headers):
            raise ValidationException("Excel 文件缺少表头")

        rows = []
        for values in raw_rows[1:]:
            row = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                row[header] = values[idx] if idx < len(values) else None
            if any(self._to_text(v) for v in row.values()):
                rows.append(row)
        return [h for h in headers if h], rows

    def _derive_template_name_from_filename(self, filename: str) -> str:
        base = (filename or "").strip()
        if "." in base:
            base = base.rsplit(".", 1)[0]
        return base.strip()[:128]

    def _build_fields_from_structured_rows(self, normalized_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        fields: List[Dict[str, Any]] = []
        used_names: set[str] = set()

        for idx, row in enumerate(normalized_rows):
            raw_field_name = self._to_text(row.get("field_name"))
            raw_display_name = self._to_text(row.get("field_display_name"))
            if not raw_field_name and not raw_display_name:
                continue

            preferred_name = raw_field_name or raw_display_name
            field_name = self._deduplicate_name(self._normalize_field_name(preferred_name, idx + 1), used_names)
            used_names.add(field_name)
            display_name = (raw_display_name or raw_field_name or f"字段{idx + 1}")[:128]

            field_type = self._normalize_field_type(row.get("field_type"))
            required = self._parse_bool(row.get("field_required"))
            description = self._to_text(row.get("field_description")) or None
            extraction_hints = self._to_text(row.get("field_extraction_hints")) or None
            sort_order = self._parse_int(row.get("field_sort_order"), len(fields))

            fields.append(
                {
                    "name": field_name,
                    "display_name": display_name,
                    "field_type": field_type,
                    "description": description,
                    "required": required,
                    "extraction_hints": extraction_hints,
                    "sort_order": sort_order,
                }
            )

        return fields

    def _build_fields_from_simple_headers(
        self,
        headers: List[str],
        row_data: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        alias_keys = {
            self._normalize_header(alias)
            for aliases in TEMPLATE_TABULAR_ALIAS.values()
            for alias in aliases
        }

        display_headers: List[str] = []
        seen: set[str] = set()
        for header in headers or []:
            display_name = self._to_text(header)
            if not display_name:
                continue
            normalized = self._normalize_header(display_name)
            if normalized in seen:
                continue
            seen.add(normalized)
            if normalized in alias_keys:
                continue
            display_headers.append(display_name)

        if not display_headers and headers:
            # 若全部命中别名（极端情况），回退使用原始非空表头。
            display_headers = [self._to_text(h) for h in headers if self._to_text(h)]

        fields: List[Dict[str, Any]] = []
        used_names: set[str] = set()
        for idx, display_name in enumerate(display_headers):
            sample_values = []
            for row in row_data or []:
                value = row.get(display_name)
                if self._to_text(value):
                    sample_values.append(value)

            field_name = self._deduplicate_name(
                self._normalize_field_name(display_name, idx + 1),
                used_names,
            )
            used_names.add(field_name)

            fields.append(
                {
                    "name": field_name,
                    "display_name": display_name[:128],
                    "field_type": self._infer_field_type_from_header_and_samples(display_name, sample_values),
                    "description": f"从表头提取{display_name}",
                    "required": False,
                    "extraction_hints": None,
                    "sort_order": idx,
                }
            )

        return fields

    def _decode_csv_text(self, file_content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
            try:
                return file_content.decode(encoding)
            except UnicodeDecodeError:
                continue

        try:
            import chardet
            detect_result = chardet.detect(file_content)
            guessed = detect_result.get("encoding")
            if guessed:
                return file_content.decode(guessed, errors="ignore")
        except Exception:
            pass

        raise ValidationException("CSV 编码无法识别，请使用 UTF-8 或 GBK 编码")

    def _normalize_template_import_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        alias_map: Dict[str, str] = {}
        for canonical, aliases in TEMPLATE_TABULAR_ALIAS.items():
            for alias in aliases:
                alias_map[self._normalize_header(alias)] = canonical

        normalized: Dict[str, Any] = {}
        for key, value in (row or {}).items():
            canonical = alias_map.get(self._normalize_header(key))
            if not canonical:
                continue
            if canonical not in normalized or not self._to_text(normalized.get(canonical)):
                normalized[canonical] = value

        return normalized

    def _normalize_header(self, value: Any) -> str:
        text = self._to_text(value).lower()
        return re.sub(r"[\s_\-]", "", text)

    def _to_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _first_non_empty(self, rows: List[Dict[str, Any]], key: str) -> str:
        for row in rows:
            value = self._to_text(row.get(key))
            if value:
                return value
        return ""

    def _parse_bool(self, value: Any) -> bool:
        if isinstance(value, bool):
            return value
        text = self._to_text(value).lower()
        if text in {"1", "true", "yes", "y", "是", "必填", "需要"}:
            return True
        if text in {"0", "false", "no", "n", "否", "非必填", "不需要", ""}:
            return False
        return False

    def _parse_int(self, value: Any, default: int = 0) -> int:
        text = self._to_text(value)
        if not text:
            return default
        try:
            return int(float(text))
        except ValueError:
            return default

    def _parse_template_status(self, value: Any) -> TemplateStatus:
        text = self._to_text(value).lower()
        mapping = {
            "draft": TemplateStatus.DRAFT,
            "草稿": TemplateStatus.DRAFT,
            "active": TemplateStatus.ACTIVE,
            "发布": TemplateStatus.ACTIVE,
            "已发布": TemplateStatus.ACTIVE,
            "deprecated": TemplateStatus.DEPRECATED,
            "废弃": TemplateStatus.DEPRECATED,
            "已废弃": TemplateStatus.DEPRECATED,
            "archived": TemplateStatus.ARCHIVED,
            "归档": TemplateStatus.ARCHIVED,
            "已归档": TemplateStatus.ARCHIVED,
        }
        return mapping.get(text, TemplateStatus.DRAFT)

    def _build_template_export_headers(self, template: Template) -> List[str]:
        fields = sorted(
            list(template.fields or []),
            key=lambda f: (f.sort_order if f.sort_order is not None else 0),
        )
        return [
            self._to_text(getattr(field, "display_name", None))
            or self._to_text(getattr(field, "name", None))
            for field in fields
            if self._to_text(getattr(field, "display_name", None))
            or self._to_text(getattr(field, "name", None))
        ]

    async def infer_template_from_document(
        self,
        document_id: str,
        requester_id: str,
        requester_is_superuser: bool,
        template_name: Optional[str] = None,
        description: Optional[str] = None,
        max_fields: int = 50,
        on_chunk_done: Optional[Callable[[Dict[str, Any]], Awaitable[None]]] = None,
    ) -> Dict[str, Any]:
        query = select(Document).where(
            Document.id == document_id,
            Document.deleted_at.is_(None),
        )
        if not requester_is_superuser:
            query = query.where(Document.owner_id == requester_id)

        result = await self.db.execute(query)
        document = result.scalar_one_or_none()
        if not document:
            raise NotFoundException("文档", document_id)
        if document.status != DocumentStatus.PROCESSED:
            raise ValidationException(f"文档尚未解析完成，当前状态: {document.status.value}")

        file_content = storage.download_file(document.storage_bucket, document.storage_path)
        mime_type = normalize_mime_type(document.mime_type, document.name)
        processor = get_processor(mime_type)
        if not processor:
            raise ValidationException(f"文档不支持解析，MIME={mime_type}")

        parse_result = await processor.parse(file_content, document.name)
        full_text = (parse_result.full_text or "").strip()
        if not full_text:
            raise ValidationException("文档内容为空，无法自动生成模板")

        chunk_size = max(1000, int(getattr(settings, "TEMPLATE_INFER_CHUNK_SIZE", TEMPLATE_INFER_CHUNK_SIZE)))
        chunk_overlap = max(0, int(getattr(settings, "TEMPLATE_INFER_CHUNK_OVERLAP", TEMPLATE_INFER_CHUNK_OVERLAP)))
        max_chunks = max(1, int(getattr(settings, "TEMPLATE_INFER_MAX_CHUNKS", TEMPLATE_INFER_MAX_CHUNKS)))
        text_chunks = self._split_text_with_overlap(full_text, chunk_size, chunk_overlap)
        selected_chunks = text_chunks[:max_chunks]
        inferred_fields: List[Dict[str, Any]] = []
        inferred_name = (template_name or "").strip() or self._default_template_name(document.name)
        inferred_desc = (description or "").strip() or f"基于文档《{document.display_name or document.name}》自动生成"

        logger.info(
            "模板自动推断开始: document_id=%s requester_id=%s mime_type=%s doc_len=%s chunk_size=%s overlap=%s total_chunks=%s selected_chunks=%s max_fields=%s",
            document_id,
            requester_id,
            mime_type,
            len(full_text),
            chunk_size,
            chunk_overlap,
            len(text_chunks),
            len(selected_chunks),
            max_fields,
        )

        # Excel 优先使用真实表头做确定性推断，避免LLM异常时退化为固定兜底字段。
        is_excel_doc = mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if is_excel_doc:
            excel_fields = self._infer_fields_from_excel_tables(parse_result, max_fields)
            if excel_fields:
                logger.info(
                    "模板自动推断命中Excel确定性分支: document_id=%s field_count=%s",
                    document_id,
                    len(excel_fields),
                )
                return {
                    "name": inferred_name,
                    "description": inferred_desc,
                    "fields": excel_fields,
                }

        try:
            adapter, llm_config = await self._resolve_infer_adapter_and_config()
            llm_config.temperature = 0.1
            llm_config.max_tokens = min(llm_config.max_tokens, 1800)

            chunk_field_cap = max(
                1,
                min(
                    int(getattr(settings, "TEMPLATE_INFER_CHUNK_MAX_FIELDS", TEMPLATE_INFER_CHUNK_MAX_FIELDS)),
                    max_fields,
                ),
            )
            candidate_field_limit = max_fields * 3
            inferred_fields = []
            chunk_total = len(selected_chunks)
            chunk_concurrency = max(
                1,
                int(getattr(settings, "TEMPLATE_INFER_CHUNK_CONCURRENCY", TEMPLATE_INFER_CHUNK_CONCURRENCY)),
            )
            logger.info(
                "模板自动推断启动异步分块执行: document_id=%s chunk_total=%s chunk_concurrency=%s",
                document_id,
                chunk_total,
                chunk_concurrency,
            )

            semaphore = asyncio.Semaphore(chunk_concurrency)
            chunk_tasks = [
                asyncio.create_task(
                    self._infer_single_chunk_fields(
                        chunk_index=chunk_idx,
                        chunk_total=chunk_total,
                        chunk_text=chunk_text,
                        inferred_name=inferred_name,
                        inferred_desc=inferred_desc,
                        chunk_field_cap=chunk_field_cap,
                        adapter=adapter,
                        llm_config=llm_config,
                        document_id=document_id,
                        semaphore=semaphore,
                    )
                )
                for chunk_idx, chunk_text in enumerate(selected_chunks, start=1)
            ]

            chunk_result_map: Dict[int, Dict[str, Any]] = {}
            processed_chunks = 0
            try:
                for done in asyncio.as_completed(chunk_tasks):
                    chunk_result = await done
                    chunk_idx = chunk_result["chunk_index"]
                    chunk_result_map[chunk_idx] = chunk_result
                    processed_chunks += 1

                    parsed = chunk_result["parsed"]
                    if chunk_idx == 1:
                        inferred_name = (parsed.get("name") or inferred_name).strip() or inferred_name
                        inferred_desc = (parsed.get("description") or inferred_desc).strip() or inferred_desc

                    chunk_fields = chunk_result["chunk_fields"]
                    inferred_fields = self._merge_inferred_field_candidates(
                        inferred_fields,
                        chunk_fields,
                        candidate_field_limit,
                    )

                    logger.info(
                        "模板自动推断分块聚合进度: document_id=%s done=%s/%s chunk=%s/%s chunk_fields=%s aggregated_fields=%s",
                        document_id,
                        processed_chunks,
                        chunk_total,
                        chunk_idx,
                        chunk_total,
                        len(chunk_fields),
                        len(inferred_fields),
                    )

                    await self._emit_infer_chunk_progress(
                        on_chunk_done,
                        {
                            "stage": "chunk_done",
                            "document_id": document_id,
                            "chunk_index": chunk_idx,
                            "chunk_total": chunk_total,
                            "processed_chunks": processed_chunks,
                            "progress_percent": round((processed_chunks / max(1, chunk_total)) * 100, 2),
                            "chunk_fields": chunk_fields,
                            "aggregated_fields": self._sanitize_inferred_fields(inferred_fields, max_fields),
                            "name": inferred_name,
                            "description": inferred_desc,
                        },
                    )
            except Exception:
                for task in chunk_tasks:
                    if not task.done():
                        task.cancel()
                raise

            logger.info(
                "模板自动推断跳过准确性校验阶段，直接使用第一轮分块聚合结果: document_id=%s aggregated_field_candidates=%s",
                document_id,
                len(inferred_fields),
            )
        except Exception as e:
            logger.exception("模板自动推断模型调用失败，直接返回失败: document_id=%s error=%s", document_id, e)
            raise ValidationException(f"模板自动推断失败: {e}")

        cleaned_fields = self._sanitize_inferred_fields(inferred_fields, max_fields)
        if not cleaned_fields:
            logger.warning("模板自动推断清洗后字段为空，直接返回失败: document_id=%s", document_id)
            raise ValidationException("模板自动推断失败：未生成有效字段")

        logger.info(
            "模板自动推断完成: document_id=%s template_name=%s field_count=%s",
            document_id,
            inferred_name,
            len(cleaned_fields),
        )

        await self._emit_infer_chunk_progress(
            on_chunk_done,
            {
                "stage": "completed",
                "document_id": document_id,
                "progress_percent": 100.0,
                "name": inferred_name,
                "description": inferred_desc,
                "fields": cleaned_fields,
            },
        )

        return {
            "name": inferred_name,
            "description": inferred_desc,
            "fields": cleaned_fields,
        }

    async def _emit_infer_chunk_progress(
        self,
        callback: Optional[Callable[[Dict[str, Any]], Awaitable[None]]],
        payload: Dict[str, Any],
    ) -> None:
        if callback is None:
            return
        try:
            await callback(payload)
        except Exception as e:
            logger.warning("模板自动推断进度回调失败: error=%s", e)

    async def _infer_single_chunk_fields(
        self,
        chunk_index: int,
        chunk_total: int,
        chunk_text: str,
        inferred_name: str,
        inferred_desc: str,
        chunk_field_cap: int,
        adapter,
        llm_config,
        document_id: str,
        semaphore: asyncio.Semaphore,
    ) -> Dict[str, Any]:
        """异步执行单个分块推断。"""
        async with semaphore:
            prompt = self._build_infer_prompt(chunk_text, inferred_name, inferred_desc, chunk_field_cap)
            logger.info(
                "模板自动推断模型调用输入(第一轮分块): document_id=%s chunk=%s/%s provider=%s model=%s prompt=%s",
                document_id,
                chunk_index,
                chunk_total,
                getattr(adapter, "provider_name", "unknown"),
                llm_config.model,
                self._truncate_log_text(prompt),
            )
            messages = [
                LLMMessage(
                    role="system",
                    content=(
                        "你是资深文档结构化工程师。请根据文档内容设计信息提取模板字段。"
                        "输出必须是严格 JSON，不要包含代码块标记或额外解释。"
                    ),
                ),
                LLMMessage(role="user", content=prompt),
            ]

            response = await adapter.chat(messages, llm_config)
            logger.info(
                "模板自动推断模型调用输出(第一轮分块): document_id=%s chunk=%s/%s total_tokens=%s content=%s",
                document_id,
                chunk_index,
                chunk_total,
                response.total_tokens,
                self._truncate_log_text(response.content),
            )

            parsed = await self._parse_or_repair_infer_response(
                adapter=adapter,
                llm_config=llm_config,
                raw_content=response.content,
                document_id=document_id,
                stage=f"第一轮分块{chunk_index}",
            )
            chunk_fields = self._sanitize_inferred_fields(parsed.get("fields") or [], chunk_field_cap)
            return {
                "chunk_index": chunk_index,
                "parsed": parsed,
                "chunk_fields": chunk_fields,
            }

    def _split_text_with_overlap(self, text: str, chunk_size: int, overlap: int) -> List[str]:
        """将长文本按固定窗口切片，并保留重叠区域降低漏字段风险。"""
        if not text:
            return [""]
        if chunk_size <= 0:
            return [text]

        overlap = max(0, min(overlap, chunk_size - 1))
        chunks: List[str] = []
        start = 0
        text_len = len(text)

        while start < text_len:
            end = min(start + chunk_size, text_len)
            chunks.append(text[start:end])
            if end >= text_len:
                break
            start = end - overlap

        return chunks

    def _build_chunk_context_text(self, chunks: List[str], max_chars: int) -> str:
        """构建带切片标记的文档上下文，供校验轮统一使用。"""
        if not chunks:
            return ""

        max_chars = max(1000, int(max_chars or TEMPLATE_INFER_VERIFY_TEXT_LIMIT))
        parts: List[str] = []
        total = 0
        chunk_total = len(chunks)
        for idx, chunk in enumerate(chunks, start=1):
            if total >= max_chars:
                break
            header = f"[文档切片 {idx}/{chunk_total}]\n"
            remain = max_chars - total - len(header)
            if remain <= 0:
                break
            body = chunk[:remain]
            segment = f"{header}{body}\n"
            parts.append(segment)
            total += len(segment)

        return "\n".join(parts)

    def _merge_inferred_field_candidates(
        self,
        existing_fields: List[Dict[str, Any]],
        incoming_fields: List[Dict[str, Any]],
        max_candidates: int,
    ) -> List[Dict[str, Any]]:
        """合并分块字段候选并去重，控制校验轮输入规模。"""
        merged = list(existing_fields or [])
        seen_names = {
            str(item.get("name") or "").strip().lower()
            for item in merged
            if isinstance(item, dict)
        }
        seen_display = {
            str(item.get("display_name") or "").strip().lower()
            for item in merged
            if isinstance(item, dict)
        }

        for item in incoming_fields or []:
            if not isinstance(item, dict):
                continue
            name_key = str(item.get("name") or "").strip().lower()
            display_key = str(item.get("display_name") or "").strip().lower()
            if name_key and name_key in seen_names:
                continue
            if display_key and display_key in seen_display:
                continue

            merged.append(item)
            if name_key:
                seen_names.add(name_key)
            if display_key:
                seen_display.add(display_key)

            if len(merged) >= max_candidates:
                break

        return merged

    def _default_template_name(self, filename: str) -> str:
        base = (filename or "未命名文档").rsplit(".", 1)[0].strip()
        return f"{base}提取模板"

    def _build_infer_prompt(self, doc_text: str, template_name: str, description: str, max_fields: int) -> str:
        allowed_types = ", ".join([t.value for t in FieldType])
        return (
            "请从以下文档内容设计一个可用于结构化提取的模板。\n"
            f"目标模板名: {template_name}\n"
            f"模板描述: {description}\n"
            f"最多输出字段数: {max_fields}\n"
            f"字段类型只能使用: {allowed_types}\n\n"
            "输出格式必须是以下 JSON 对象:\n"
            "{\n"
            "  \"name\": \"模板名称\",\n"
            "  \"description\": \"模板描述\",\n"
            "  \"fields\": [\n"
            "    {\n"
            "      \"name\": \"english_or_zh_key\",\n"
            "      \"display_name\": \"字段显示名\",\n"
            "      \"field_type\": \"text\",\n"
            "      \"description\": \"字段含义\",\n"
            "      \"required\": false,\n"
            "      \"extraction_hints\": \"提取提示\"\n"
            "    }\n"
            "  ]\n"
            "}\n\n"
            "要求:\n"
            "1) 字段名必须可读且唯一。\n"
            "2) 只输出文档中可能稳定出现的关键字段。\n"
            "3) 不要输出任何 JSON 之外内容。\n\n"
            "文档内容如下:\n"
            f"{doc_text}"
        )

    def _build_verify_prompt(
        self,
        doc_text: str,
        template_name: str,
        description: str,
        fields: List[Dict[str, Any]],
        max_fields: int,
    ) -> str:
        allowed_types = ", ".join([t.value for t in FieldType])
        return (
            "请对以下模板字段进行准确性校验，并输出修正后的字段列表。\n"
            "校验要求:\n"
            "1) 删除无法从文档稳定提取的字段；\n"
            "2) 修正字段类型与字段名；\n"
            "3) 字段名必须唯一且可读；\n"
            f"4) 最多保留 {max_fields} 个字段；\n"
            f"5) 字段类型只能使用: {allowed_types}。\n\n"
            f"模板名: {template_name}\n"
            f"模板描述: {description}\n"
            f"候选字段: {json.dumps(fields, ensure_ascii=False)}\n\n"
            "输出 JSON 格式:\n"
            "{\n"
            "  \"name\": \"模板名称\",\n"
            "  \"description\": \"模板描述\",\n"
            "  \"fields\": [\n"
            "    {\n"
            "      \"name\": \"field_key\",\n"
            "      \"display_name\": \"字段显示名\",\n"
            "      \"field_type\": \"text\",\n"
            "      \"description\": \"字段含义\",\n"
            "      \"required\": false,\n"
            "      \"extraction_hints\": \"提取提示\"\n"
            "    }\n"
            "  ]\n"
            "}\n\n"
            "文档内容如下:\n"
            f"{doc_text}"
        )

    def _parse_infer_response(self, content: str) -> Dict[str, Any]:
        text = (content or "").strip()
        if not text:
            raise ValidationException("LLM 返回内容为空，无法解析为 JSON")

        if text.startswith("```"):
            lines = text.splitlines()
            if len(lines) >= 3:
                text = "\n".join(lines[1:-1]).strip()

        candidates: List[str] = [text]

        # 尝试提取最外层 JSON 对象，避免模型前后夹杂解释文本。
        first_brace = text.find("{")
        last_brace = text.rfind("}")
        if first_brace != -1 and last_brace != -1 and first_brace < last_brace:
            candidates.append(text[first_brace:last_brace + 1])

        for candidate in candidates:
            parsed = self._try_parse_json_candidate(candidate)
            if isinstance(parsed, dict):
                return parsed

        raise ValidationException("LLM 返回内容无法解析为 JSON")

    async def _parse_or_repair_infer_response(
        self,
        adapter,
        llm_config,
        raw_content: str,
        document_id: str,
        stage: str,
    ) -> Dict[str, Any]:
        """先本地解析；失败时让模型仅做 JSON 修复重写。"""
        try:
            return self._parse_infer_response(raw_content)
        except Exception as parse_err:
            logger.warning(
                "模板自动推断%s解析失败，尝试JSON修复重写: document_id=%s error=%s",
                stage,
                document_id,
                parse_err,
            )

        repair_messages = [
            LLMMessage(
                role="system",
                content=(
                    "你是JSON修复器。把用户给出的文本重写为严格JSON对象。"
                    "只输出JSON对象本身，不要任何解释。"
                    "必须包含键: name, description, fields。"
                ),
            ),
            LLMMessage(
                role="user",
                content=(
                    "请将下面内容修复为严格JSON。"
                    "如果原文包含多余说明、注释、代码块、单引号、尾逗号，请清理。\n\n"
                    f"原始内容:\n{self._truncate_log_text(raw_content, 12000)}"
                ),
            ),
        ]

        repair_response = await adapter.chat(repair_messages, llm_config)
        logger.info(
            "模板自动推断%s JSON修复输出: document_id=%s total_tokens=%s content=%s",
            stage,
            document_id,
            repair_response.total_tokens,
            self._truncate_log_text(repair_response.content),
        )

        return self._parse_infer_response(repair_response.content)

    def _try_parse_json_candidate(self, text: str) -> Optional[Dict[str, Any]]:
        """对候选文本做容错 JSON 解析。"""
        if not text:
            return None

        raw = text.strip()
        attempts = [raw]

        # 常见修复1：去掉对象/数组前的尾逗号。
        attempts.append(re.sub(r",\s*([}\]])", r"\1", raw))

        # 常见修复2：处理Python字面量风格（单引号/True/False/None）。
        normalized = re.sub(r",\s*([}\]])", r"\1", raw)
        try:
            literal_obj = ast.literal_eval(normalized)
            if isinstance(literal_obj, dict):
                return literal_obj
        except Exception:
            pass

        for attempt in attempts:
            try:
                parsed = json.loads(attempt)
                if isinstance(parsed, dict):
                    return parsed
            except Exception:
                continue

        return None

    def _sanitize_inferred_fields(self, fields: List[Dict[str, Any]], max_fields: int) -> List[Dict[str, Any]]:
        used_names = set()
        cleaned = []
        for idx, raw in enumerate(fields[:max_fields]):
            if not isinstance(raw, dict):
                continue

            name = self._normalize_field_name(raw.get("name"), idx + 1)
            name = self._deduplicate_name(name, used_names)
            used_names.add(name)

            display_name = str(raw.get("display_name") or raw.get("name") or f"字段{idx + 1}").strip()
            field_type = self._normalize_field_type(raw.get("field_type"))

            cleaned.append(
                {
                    "name": name,
                    "display_name": display_name[:128],
                    "field_type": field_type,
                    "description": (str(raw.get("description") or "").strip() or None),
                    "required": bool(raw.get("required", False)),
                    "extraction_hints": (str(raw.get("extraction_hints") or "").strip() or None),
                    "sort_order": idx,
                }
            )
        return cleaned

    def _normalize_field_name(self, value: Any, idx: int) -> str:
        text = str(value or "").strip()
        if not text:
            return f"field_{idx}"

        text = re.sub(r"\s+", "_", text)
        text = re.sub(r"[^\u4e00-\u9fff_a-zA-Z0-9]", "_", text)
        text = re.sub(r"_+", "_", text).strip("_")
        if not text:
            return f"field_{idx}"
        if re.match(r"^[0-9]", text):
            text = f"field_{text}"
        return text[:64]

    def _deduplicate_name(self, name: str, used_names: set[str]) -> str:
        if name not in used_names:
            return name
        suffix = 2
        while f"{name}_{suffix}" in used_names:
            suffix += 1
        return f"{name}_{suffix}"[:64]

    def _normalize_field_type(self, value: Any) -> str:
        raw = str(value or "").strip().lower()
        alias = {
            "integer": "number",
            "float": "number",
            "double": "number",
            "money": "number",
            "amount": "number",
            "time": "datetime",
            "bool": "boolean",
            "array": "list",
            "object": "custom",
        }
        normalized = alias.get(raw, raw)
        if normalized in {t.value for t in FieldType}:
            return normalized
        return FieldType.TEXT.value

    def _fallback_fields_from_text(self, text: str, max_fields: int) -> List[Dict[str, Any]]:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        candidate_names = []
        for line in lines[:300]:
            if "：" in line:
                key = line.split("：", 1)[0].strip()
            elif ":" in line:
                key = line.split(":", 1)[0].strip()
            else:
                continue
            if 1 <= len(key) <= 20:
                candidate_names.append(key)

        unique = []
        seen = set()
        for name in candidate_names:
            if name in seen:
                continue
            seen.add(name)
            unique.append(name)
            if len(unique) >= max_fields:
                break

        if not unique:
            unique = ["编号", "名称", "日期", "金额"][:max_fields]

        fields = []
        used_names = set()
        for idx, display_name in enumerate(unique):
            key = self._deduplicate_name(self._normalize_field_name(display_name, idx + 1), used_names)
            used_names.add(key)
            field_type = FieldType.NUMBER.value if any(k in display_name for k in ["金额", "数量", "总计"]) else FieldType.TEXT.value
            if any(k in display_name for k in ["日期", "时间"]):
                field_type = FieldType.DATE.value
            fields.append(
                {
                    "name": key,
                    "display_name": display_name,
                    "field_type": field_type,
                    "description": f"从文档中提取{display_name}",
                    "required": False,
                    "extraction_hints": None,
                    "sort_order": idx,
                }
            )
        return fields

    def _infer_fields_from_excel_tables(self, parse_result, max_fields: int) -> List[Dict[str, Any]]:
        """从 Excel 解析结果中直接提取表头并推断字段类型。"""
        if max_fields <= 0:
            return []

        header_samples: List[Tuple[str, List[Any]]] = []
        seen_headers = set()

        for page in parse_result.pages or []:
            table = (page.tables or [None])[0]
            if not isinstance(table, dict):
                continue

            headers = table.get("headers") if isinstance(table.get("headers"), list) else []
            rows = table.get("rows") if isinstance(table.get("rows"), list) else []
            if not headers:
                continue

            col_count = max(len(headers), max((len(r) for r in rows), default=0))
            for col_idx in range(col_count):
                raw_header = headers[col_idx] if col_idx < len(headers) else ""
                header = str(raw_header).strip() if raw_header is not None else ""
                if not header:
                    continue

                # 跳过明显占位字段名
                if re.match(r"^column_\d+$", header, flags=re.IGNORECASE):
                    continue

                header_key = header.lower()
                if header_key in seen_headers:
                    continue
                seen_headers.add(header_key)

                samples = []
                for row in rows[:60]:
                    if not isinstance(row, list) or col_idx >= len(row):
                        continue
                    value = row[col_idx]
                    if value in (None, ""):
                        continue
                    samples.append(value)

                header_samples.append((header, samples))
                if len(header_samples) >= max_fields:
                    break

            if len(header_samples) >= max_fields:
                break

        if not header_samples:
            return []

        fields: List[Dict[str, Any]] = []
        used_names = set()
        for idx, (display_name, samples) in enumerate(header_samples):
            key = self._deduplicate_name(self._normalize_field_name(display_name, idx + 1), used_names)
            used_names.add(key)
            field_type = self._infer_field_type_from_header_and_samples(display_name, samples)

            fields.append(
                {
                    "name": key,
                    "display_name": display_name[:128],
                    "field_type": field_type,
                    "description": f"从Excel表头提取{display_name}",
                    "required": False,
                    "extraction_hints": None,
                    "sort_order": idx,
                }
            )

        return fields

    def _infer_field_type_from_header_and_samples(self, header: str, samples: List[Any]) -> str:
        """基于表头语义和样本值推断字段类型。"""
        header_l = (header or "").strip().lower()

        if any(k in header_l for k in ["date", "time", "日期", "时间"]):
            return FieldType.DATE.value
        if any(k in header_l for k in ["amount", "price", "total", "qty", "count", "num", "金额", "价格", "数量", "总计"]):
            return FieldType.NUMBER.value

        non_empty = [str(v).strip() for v in samples if v not in (None, "")]
        if not non_empty:
            return FieldType.TEXT.value

        num_count = 0
        date_count = 0
        for v in non_empty:
            if re.match(r"^-?\d+(\.\d+)?$", v.replace(",", "")):
                num_count += 1
            if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", v):
                date_count += 1

        total = len(non_empty)
        if date_count / total >= 0.6:
            return FieldType.DATE.value
        if num_count / total >= 0.7:
            return FieldType.NUMBER.value
        return FieldType.TEXT.value
